"""
Módulo responsável pela extração e processamento de dados dos arquivos Excel.
Utiliza mapeamentos de constantes e mensagens centralizadas.
"""

import logging
from os import path
import openpyxl
from pathlib import Path
from package import COMPANIES_REGISTRY, DEPARTMENTS_CONFIG
from package import get_msg
from package import setup_logging

logger = logging.getLogger("summarize-budget")


def process_budget_data(target_company_name: str, dept_code: str):
    """
    Lê os dados de orçamento das colunas D a O e gera o arquivo de saída formatado.
    """

    # 1. Obter configurações do registro
    company = COMPANIES_REGISTRY.get(target_company_name)
    if not company:
        logger.warning(f"Empresa '{target_company_name}' não encontrada no COMPANIES_REGISTRY.")
        logger.warning(get_msg("ERR_COMP_NOT_FOUND", add_info=target_company_name))
        return

    # Localiza o input específico para o internal_dept_code fornecido
    input_cfg = next((i for i in company.inputs if i.internal_dept_code == dept_code), None)
    if not input_cfg:
        # logger.warning(f"Departamento '{dept_code}' não configurado para a empresa {target_company_name}.")
        logger.warning(get_msg("ERR_DEPT_NOT_CONFIGURED", dept_code=dept_code, company=target_company_name))
        return

    # 2. Configurações de Caminho e Variáveis Fixas
    # Conforme sua preferência: bases/Contorno para as entradas
    input_folder = Path("bases/Contorno")
    input_path = input_folder / input_cfg.file_name
    output_folder = Path("output")
    output_folder.mkdir(exist_ok=True)

    target_year = 2026
    origin_fixed = "Orçamento"
    historic_text = ""
    view_fixed = "Atual"

    if not input_path.exists():
        # logger.warning(f"Erro: ERR_FILE_NOT_FOUND" {input_path}")
        logger.warning(get_msg("ERR_FILE_NOT_FOUND", add_info=input_path))
        return

    try:
        # Carrega o workbook (data_only=True para ler o resultado de fórmulas)
        wb = openpyxl.load_workbook(input_path, data_only=True)
        if input_cfg.sheet_name not in wb.sheetnames:
            logger.warning(get_msg("ERR_SHEET_NOT_FOUND", path=input_cfg.sheet_name))
            return

        sheet = wb[input_cfg.sheet_name]
        final_rows = []

        # 3. Processamento das linhas conforme DEPARTMENTS_CONFIG
        dept_mapping = DEPARTMENTS_CONFIG.get(dept_code, {})

        for excel_line, account_entry in dept_mapping.items():
            # Loop de Janeiro (Mês 1) a Dezembro (Mês 12)
            # Conforme sua informação: Coluna D (4) até O (15)
            for month_idx in range(1, 13):
                # Cálculo da coluna: Jan=4(D), Fev=5(E), ..., Dez=15(O)
                col_idx = month_idx + 3 

                raw_value = sheet.cell(row=excel_line, column=col_idx).value

                # Tratamento de valores nulos e conversão
                if raw_value is None or raw_value == "":
                    value_float = 0.0
                else:
                    try:
                        value_float = float(raw_value)
                    except ValueError:
                        value_float = 0.0

                # Aplica o fator da conta e arredonda para 2 casas decimais
                clean_value = round(value_float * account_entry.factor, 2)

                # Monta o dicionário conforme a ordem das colunas desejada
                row_data = {
                    "emp": company.id,
                    "mes": month_idx,
                    "ano": target_year,
                    "origem": origin_fixed,
                    "historico": historic_text,
                    "centro_custo_bal": input_cfg.accounting_dept_name,
                    "conta": account_entry.account_code,
                    "des_conta": account_entry.description,
                    "movimento": clean_value,
                    "view": view_fixed
                }
                final_rows.append(row_data)

        # 4. Criação do arquivo de saída Excel
        output_filename = f"output_{target_company_name.replace(' ', '_')}_{dept_code}.xlsx"
        output_path = output_folder / output_filename

        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.title = "Orcamento_Processado"

        columns_order = [
            "emp", "mes", "ano", "origem", "historico",
            "centro_custo_bal", "conta", "des_conta", "movimento", "view"
        ]

        # Escreve o cabeçalho
        output_sheet.append(columns_order)

        # Escreve os dados processados
        for data in final_rows:
            output_sheet.append([data[col] for col in columns_order])

        output_wb.save(output_path)
        logger.info(get_msg("INFO_PROCESS_SUCCESS", add_info=output_path))
        logger.info(get_msg("INFO_FILE_GENERATED", add_info=output_path))

    except Exception as e:
        logger.error(f"Ocorreu um erro durante o processamento: {e}")
